"""
סידור עבודה — Vercel + Supabase PostgreSQL
"""

import os, bcrypt, psycopg2, psycopg2.extras
import jwt as pyjwt
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from functools import wraps
from flask import Flask, request, jsonify

app = Flask(__name__)

JWT_SECRET = os.environ.get('JWT_SECRET', 'change-me-in-vercel')
JWT_EXPIRY_HOURS = 24

def create_token(payload):
    data = dict(payload)
    data['exp'] = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS)
    return pyjwt.encode(data, JWT_SECRET, algorithm='HS256')

def verify_token():
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        raise Exception('no token')
    token = auth.split(' ', 1)[1]
    return pyjwt.decode(token, JWT_SECRET, algorithms=['HS256'])

DOW_NAMES = ['ראשון','שני','שלישי','רביעי','חמישי','שישי','שבת']

def date_to_dow(date_str):
    return datetime.strptime(date_str, '%Y-%m-%d').isoweekday() % 7  # 0=Sun…6=Sat

@contextmanager
def get_db():
    conn = psycopg2.connect(
        os.environ['DATABASE_URL'],
        cursor_factory=psycopg2.extras.RealDictCursor,
        sslmode='require'
    )
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    traceback.print_exc()
    return jsonify(error=str(e)), 500

def init_db():
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin','lecturer'))
            );
            CREATE TABLE IF NOT EXISTS courses (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS lecturer_courses (
                lecturer_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                course_id   INTEGER NOT NULL REFERENCES courses(id) ON DELETE CASCADE,
                PRIMARY KEY (lecturer_id, course_id)
            );
            CREATE TABLE IF NOT EXISTS slots (
                id SERIAL PRIMARY KEY,
                date TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                course_id INTEGER NOT NULL REFERENCES courses(id) ON DELETE CASCADE,
                status TEXT DEFAULT 'open' CHECK(status IN ('open','filled')),
                assigned_lecturer_id INTEGER REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS lecturer_availability (
                id SERIAL PRIMARY KEY,
                lecturer_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                type TEXT NOT NULL CHECK(type IN ('weekly','date')),
                day_of_week INTEGER,
                date TEXT,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                is_available INTEGER DEFAULT 1
            );
        """)
        cur.execute("SELECT id FROM users WHERE role='admin' LIMIT 1")
        if not cur.fetchone():
            pw = bcrypt.hashpw(b'admin123', bcrypt.gensalt()).decode()
            cur.execute(
                "INSERT INTO users(name,username,password_hash,role) VALUES(%s,%s,%s,'admin')",
                ('מנהל', 'admin', pw)
            )

try:
    init_db()
except Exception as e:
    print(f"DB init: {e}")

# ── Auth ──────────────────────────────────────────────────────────────────────

def require_auth(role=None):
    def dec(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            try:
                user = verify_token()
            except Exception:
                return jsonify(error='לא מחובר'), 401
            if role and user.get('role') != role:
                return jsonify(error='אין הרשאה'), 403
            return f(*args, user=user, **kwargs)
        return wrapper
    return dec

@app.route('/api/login', methods=['POST'])
def login():
    d = request.json or {}
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username=%s", (d.get('username',''),))
        user = cur.fetchone()
    if not user: return jsonify(error='שם משתמש או סיסמה שגויים'), 401
    stored = user['password_hash']
    if isinstance(stored, str): stored = stored.encode()
    if not bcrypt.checkpw(d.get('password','').encode(), stored):
        return jsonify(error='שם משתמש או סיסמה שגויים'), 401
    token = create_token({'id':user['id'],'name':user['name'],'username':user['username'],'role':user['role']})
    return jsonify(role=user['role'], name=user['name'], token=token)

@app.route('/api/logout', methods=['POST'])
def logout():
    return jsonify(ok=True)

@app.route('/api/me')
def me():
    try:
        u = verify_token()
        return jsonify(id=u['id'], name=u['name'], role=u['role'])
    except Exception:
        return jsonify(error='לא מחובר'), 401

# ── Admin: Courses ────────────────────────────────────────────────────────────

@app.route('/api/admin/courses', methods=['GET'])
@require_auth('admin')
def get_courses(user):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM courses ORDER BY name")
        return jsonify([dict(r) for r in cur.fetchall()])

@app.route('/api/admin/courses', methods=['POST'])
@require_auth('admin')
def add_course(user):
    name = (request.json or {}).get('name','').strip()
    if not name: return jsonify(error='שם חסר'), 400
    try:
        with get_db() as conn:
            conn.cursor().execute("INSERT INTO courses(name) VALUES(%s)", (name,))
        return jsonify(ok=True)
    except psycopg2.IntegrityError:
        return jsonify(error='קורס כבר קיים'), 409

@app.route('/api/admin/courses/<int:cid>', methods=['DELETE'])
@require_auth('admin')
def delete_course(cid, user):
    with get_db() as conn:
        conn.cursor().execute("DELETE FROM courses WHERE id=%s", (cid,))
    return jsonify(ok=True)

# ── Admin: Lecturers ──────────────────────────────────────────────────────────

@app.route('/api/admin/lecturers', methods=['GET'])
@require_auth('admin')
def get_lecturers(user):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id,name,username FROM users WHERE role='lecturer' ORDER BY name")
        lecs = [dict(r) for r in cur.fetchall()]
        for l in lecs:
            cur.execute("SELECT c.id,c.name FROM courses c JOIN lecturer_courses lc ON lc.course_id=c.id WHERE lc.lecturer_id=%s", (l['id'],))
            l['courses'] = [dict(r) for r in cur.fetchall()]
    return jsonify(lecs)

@app.route('/api/admin/lecturers', methods=['POST'])
@require_auth('admin')
def add_lecturer(user):
    d = request.json or {}
    name, username, password = d.get('name',''), d.get('username',''), d.get('password','')
    course_ids = d.get('courseIds', [])
    if not all([name,username,password]): return jsonify(error='שדות חסרים'), 400
    try:
        with get_db() as conn:
            cur = conn.cursor()
            pw = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
            cur.execute("INSERT INTO users(name,username,password_hash,role) VALUES(%s,%s,%s,'lecturer') RETURNING id", (name,username,pw))
            lid = cur.fetchone()['id']
            for cid in course_ids:
                cur.execute("INSERT INTO lecturer_courses(lecturer_id,course_id) VALUES(%s,%s) ON CONFLICT DO NOTHING", (lid,cid))
        return jsonify(ok=True)
    except psycopg2.IntegrityError:
        return jsonify(error='שם משתמש כבר קיים'), 409

@app.route('/api/admin/lecturers/<int:lid>', methods=['DELETE'])
@require_auth('admin')
def delete_lecturer(lid, user):
    with get_db() as conn:
        conn.cursor().execute("DELETE FROM users WHERE id=%s AND role='lecturer'", (lid,))
    return jsonify(ok=True)

@app.route('/api/admin/lecturers/<int:lid>/courses', methods=['PUT'])
@require_auth('admin')
def update_lecturer_courses(lid, user):
    course_ids = (request.json or {}).get('courseIds', [])
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM lecturer_courses WHERE lecturer_id=%s", (lid,))
        for cid in course_ids:
            cur.execute("INSERT INTO lecturer_courses(lecturer_id,course_id) VALUES(%s,%s) ON CONFLICT DO NOTHING", (lid,cid))
    return jsonify(ok=True)

# ── Admin: Slots ──────────────────────────────────────────────────────────────

@app.route('/api/admin/slots', methods=['GET'])
@require_auth('admin')
def get_slots(user):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT s.*,c.name as course_name,u.name as assigned_lecturer_name
            FROM slots s JOIN courses c ON c.id=s.course_id
            LEFT JOIN users u ON u.id=s.assigned_lecturer_id
            ORDER BY s.date,s.start_time""")
        return jsonify([dict(r) for r in cur.fetchall()])

@app.route('/api/admin/slots', methods=['POST'])
@require_auth('admin')
def add_slot(user):
    d = request.json or {}
    date,start,end,cid = d.get('date'),d.get('start_time'),d.get('end_time'),d.get('course_id')
    if not all([date,start,end,cid]): return jsonify(error='שדות חסרים'), 400
    with get_db() as conn:
        conn.cursor().execute("INSERT INTO slots(date,start_time,end_time,course_id) VALUES(%s,%s,%s,%s)", (date,start,end,cid))
    return jsonify(ok=True)

@app.route('/api/admin/slots/<int:sid>', methods=['DELETE'])
@require_auth('admin')
def delete_slot(sid, user):
    with get_db() as conn:
        conn.cursor().execute("DELETE FROM slots WHERE id=%s", (sid,))
    return jsonify(ok=True)

@app.route('/api/admin/slots/<int:sid>/assign', methods=['POST'])
@require_auth('admin')
def assign_slot(sid, user):
    lid = (request.json or {}).get('lecturer_id')
    if not lid: return jsonify(error='חסר lecturer_id'), 400
    with get_db() as conn:
        conn.cursor().execute("UPDATE slots SET assigned_lecturer_id=%s,status='filled' WHERE id=%s", (lid,sid))
    return jsonify(ok=True)

@app.route('/api/admin/slots/<int:sid>/unassign', methods=['POST'])
@require_auth('admin')
def unassign_slot(sid, user):
    with get_db() as conn:
        conn.cursor().execute("UPDATE slots SET assigned_lecturer_id=NULL,status='open' WHERE id=%s", (sid,))
    return jsonify(ok=True)

# ── Admin: Excel Download Template ───────────────────────────────────────────

@app.route('/api/admin/slots/template', methods=['GET'])
@require_auth('admin')
def download_slots_template(user):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'משבצות'
    ws.sheet_view.rightToLeft = True
    headers = ['תאריך (DD/MM/YYYY)', 'שעת התחלה (HH:MM)', 'שעת סיום (HH:MM)', 'שם קורס']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='F9D811')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.append(['01/09/2025', '08:00', '10:00', 'בישול בסיסי'])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='slots_template.xlsx')

# ── Admin: Excel Upload Slots ─────────────────────────────────────────────────

@app.route('/api/admin/slots/upload', methods=['POST'])
@require_auth('admin')
def upload_slots(user):
    from openpyxl import load_workbook
    from io import BytesIO
    f = request.files.get('file')
    if not f:
        return jsonify(error='לא נשלח קובץ'), 400
    try:
        wb = load_workbook(BytesIO(f.read()))
        ws = wb.active
    except Exception as e:
        return jsonify(error=f'שגיאה בקריאת הקובץ: {e}'), 400

    added = 0
    skipped = 0
    errors = []

    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id, name FROM courses")
        courses_map = {r['name']: r['id'] for r in cur.fetchall()}

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                date_raw, start_time, end_time, course_name = row[0], row[1], row[2], row[3]
                if not all([date_raw, start_time, end_time, course_name]):
                    skipped += 1
                    continue
                # Parse date DD/MM/YYYY → YYYY-MM-DD
                date_str = str(date_raw).strip()
                if '/' in date_str:
                    parts = date_str.split('/')
                    date_iso = f'{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}'
                else:
                    date_iso = date_str
                start_str = str(start_time).strip()[:5]
                end_str = str(end_time).strip()[:5]
                course_name_str = str(course_name).strip()
                course_id = courses_map.get(course_name_str)
                if not course_id:
                    errors.append(f'שורה {i}: קורס לא נמצא — {course_name_str}')
                    skipped += 1
                    continue
                cur.execute(
                    "INSERT INTO slots(date,start_time,end_time,course_id) VALUES(%s,%s,%s,%s)",
                    (date_iso, start_str, end_str, course_id)
                )
                added += 1
            except Exception as e:
                errors.append(f'שורה {i}: {e}')
                skipped += 1

    return jsonify(added=added, skipped=skipped, errors=errors)

# ── Admin: Suggestions ────────────────────────────────────────────────────────

def check_avail(conn, lecturer_id, slot_date, slot_start, slot_end):
    dow = date_to_dow(slot_date)
    cur = conn.cursor()
    cur.execute("SELECT * FROM lecturer_availability WHERE lecturer_id=%s AND type='date' AND date=%s", (lecturer_id,slot_date))
    date_rows = cur.fetchall()
    if date_rows:
        for e in date_rows:
            if not e['is_available']: return False
            if e['start_time'] <= slot_start and e['end_time'] >= slot_end: return True
        return False
    cur.execute("SELECT * FROM lecturer_availability WHERE lecturer_id=%s AND type='weekly' AND day_of_week=%s AND is_available=1", (lecturer_id,dow))
    for e in cur.fetchall():
        if e['start_time'] <= slot_start and e['end_time'] >= slot_end: return True
    return False

@app.route('/api/admin/suggestions', methods=['GET'])
@require_auth('admin')
def get_suggestions(user):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT s.*,c.name as course_name,u.name as assigned_lecturer_name
            FROM slots s JOIN courses c ON c.id=s.course_id
            LEFT JOIN users u ON u.id=s.assigned_lecturer_id
            ORDER BY s.date,s.start_time""")
        slots = cur.fetchall()
        result = []
        for slot in slots:
            sd = dict(slot)
            cur.execute("SELECT u.id,u.name FROM users u JOIN lecturer_courses lc ON lc.lecturer_id=u.id WHERE lc.course_id=%s AND u.role='lecturer' ORDER BY u.name", (slot['course_id'],))
            suggestions = [{'id':l['id'],'name':l['name'],'available':check_avail(conn,l['id'],slot['date'],slot['start_time'],slot['end_time'])} for l in cur.fetchall()]
            sd['suggestions'] = suggestions
            result.append(sd)
    return jsonify(result)

# ── Lecturer ──────────────────────────────────────────────────────────────────

@app.route('/api/lecturer/availability', methods=['GET'])
@require_auth('lecturer')
def get_my_avail(user):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM lecturer_availability WHERE lecturer_id=%s ORDER BY type DESC,day_of_week,date,start_time", (user['id'],))
        return jsonify([dict(r) for r in cur.fetchall()])

@app.route('/api/lecturer/availability', methods=['POST'])
@require_auth('lecturer')
def add_my_avail(user):
    d = request.json or {}
    t, start, end = d.get('type'), d.get('start_time'), d.get('end_time')
    is_av = d.get('is_available', 1)
    if t not in ('weekly','date') or not start or not end: return jsonify(error='שדות חסרים'), 400
    with get_db() as conn:
        cur = conn.cursor()
        if t == 'weekly':
            dow = d.get('day_of_week')
            if dow is None: return jsonify(error='חסר יום'), 400
            cur.execute("INSERT INTO lecturer_availability(lecturer_id,type,day_of_week,start_time,end_time,is_available) VALUES(%s,%s,%s,%s,%s,%s)", (user['id'],'weekly',dow,start,end,is_av))
        else:
            date = d.get('date')
            if not date: return jsonify(error='חסר תאריך'), 400
            cur.execute("INSERT INTO lecturer_availability(lecturer_id,type,date,start_time,end_time,is_available) VALUES(%s,%s,%s,%s,%s,%s)", (user['id'],'date',date,start,end,is_av))
    return jsonify(ok=True)

@app.route('/api/lecturer/availability/<int:aid>', methods=['DELETE'])
@require_auth('lecturer')
def del_my_avail(aid, user):
    with get_db() as conn:
        conn.cursor().execute("DELETE FROM lecturer_availability WHERE id=%s AND lecturer_id=%s", (aid,user['id']))
    return jsonify(ok=True)

@app.route('/api/lecturer/slots', methods=['GET'])
@require_auth('lecturer')
def lecturer_slots(user):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT s.id,s.date,s.start_time,s.end_time,s.status,s.assigned_lecturer_id,
            c.name as course_name,u.name as assigned_lecturer_name
            FROM slots s JOIN courses c ON c.id=s.course_id
            LEFT JOIN users u ON u.id=s.assigned_lecturer_id
            WHERE s.course_id IN (SELECT course_id FROM lecturer_courses WHERE lecturer_id=%s)
            ORDER BY s.date,s.start_time""", (user['id'],))
        return jsonify([dict(r) for r in cur.fetchall()])

# ── Logo ──────────────────────────────────────────────────────────────────────

@app.route('/danon-logo.png')
def serve_logo():
    # Serve actual PNG if present, otherwise a branded SVG placeholder
    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'danon-logo.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            return f.read(), 200, {'Content-Type': 'image/png'}
    svg = (
        '<svg xmlns="http://www.w3.org/2000/svg" width="240" height="80" viewBox="0 0 240 80">'
        '<rect width="240" height="80" rx="10" fill="#F9D811"/>'
        '<text x="120" y="54" font-family="Arial,sans-serif" font-size="36" font-weight="bold" '
        'text-anchor="middle" fill="#000">דנון</text>'
        '</svg>'
    )
    return svg, 200, {'Content-Type': 'image/svg+xml; charset=utf-8'}

# ── Serve HTML for all non-API routes ─────────────────────────────────────────

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def catch_all(path):
    html_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'index.html')
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return f.read(), 200, {'Content-Type': 'text/html; charset=utf-8'}
    except FileNotFoundError:
        return 'index.html not found', 404
